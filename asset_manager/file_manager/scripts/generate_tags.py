import openpyxl
import os

from django.conf import settings
from file_manager.models import Tag, TagGroup

os.chdir(settings.BASE_DIR + '/file_manager/scripts/tag_taxonomy/')
wb = openpyxl.load_workbook('tag_taxonomy.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

def create_tag_group(name):
    tag_group = TagGroup(name=name)
    tag_group.save()
    print('Tag Group {} created.'.format(name))
    return tag_group

def create_tag(name, tag_group):
    tag = Tag(name=name, group=tag_group)
    tag.save()
    print('Tag {0} created in Tag Group {1}.'.format(name, tag_group))

def iterate_over_column(column, row, tag_group):
    if sheet.cell(column=column, row=row).value:
        create_tag(sheet.cell(column=column, row=row).value, tag_group)
        iterate_over_column(column, row+1, tag_group)

def create_tags_by_column(column):
    tag_group = create_tag_group(sheet.cell(column=column, row=1).value)
    iterate_over_column(column, 2, tag_group)
    print()

def iterate_over_row(row, column):
    if sheet.cell(row=row, column=column).value is not None:
        create_tags_by_column(column)
        iterate_over_row(row, column+1)

def run():
    print(os.getcwd())

    iterate_over_row(1,1)


# def run():
    # print('foo')
