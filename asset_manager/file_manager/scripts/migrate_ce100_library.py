import dateutil.parser
import openpyxl
import os
import os.path

from django.conf import settings
from django.core.exceptions import ObjectDoesNotExist
from django.core.files import File

from file_manager.models import Asset, Folder, Collection, Contributor, Tag

import boto3
import botocore
s3 = boto3.resource('s3')

SOURCE_BUCKET_NAME = 's3-ce100-library'

os.chdir(settings.BASE_DIR + '/file_manager/scripts/ce100_migration/')
wb = openpyxl.load_workbook('ce100_insight_list_2017_12_05.xlsx')
insight_sheet = wb.get_sheet_by_name('result')

wb2 = openpyxl.load_workbook('ce100_tag_insight_list_2017_12_04.xlsx')
tag_sheet = wb2.get_sheet_by_name('result')

# A reference to all assets created, so tags can later be added.
asset_dict = {}

type_dict = {
    'CASE STUDY' : 'CS',
    'CO.PROJECT' : 'CP',
    'IMAGE' : 'IM',
    'LINK' : 'LN',
    'PAPER' : 'PA',
    'PRESENTATION' : 'PR',
    'REPORT' : 'RT',
    'VIDEO' : 'VI',
    'WORKSHOP SUMMARY' : 'WS'
}

def create_contributor(contributor_name):
    '''
    Creates a contributor for a given name.
    '''
    contributor = Contributor(name = contributor_name)
    contributor.save()
    print('Contributor {} created...'.format(contributor_name))
    return contributor

def get_or_create_contributors(authors):
    '''
    Returns a list of contributors from the insight authors field.
    If contributors do not yet exist, they will first be created.
    '''
    # For each contributor, get if exists, create otherwise, and return
    contributor_names = authors.split(',')
    contributor_names = [ contributor_name.strip() for contributor_name in contributor_names ]
    # eg ['foo', 'bar']

    contributors = []
    # to return all contributor models

    for contributor_name in contributor_names:
        contributor = Contributor.objects.filter(name__iexact=contributor_name).first()
        if contributor is None:
            # coontributor does not exist
            contributors.append(create_contributor(contributor_name))
        else:
            print('Contributor {} exists already...'.format(contributor_name))
            contributors.append(contributor)

    return contributors

def get_tag(tag_name):
    '''
    Returns a tag for a given name, or -1 if the tag does not exist.
    '''
    tag = Tag.objects.filter(name__iexact=tag_name).first()
    if tag is None:
        # tag does not exist
        return -1
    return tag

def get_collection_from_resource_flag(resource_flag):
    '''
    Returns the correct collection, ce100_insights or ce100_resources,
    based on the insight resource boolean.
    '''
    if resource_flag == True:
        return Collection.objects.get(id=2)
    return Collection.objects.get(id=1)

def get_filename(url):
    '''
    Returns the filename from the insight url.
    '''
    return url.rsplit('/')[-1]

def get_folder_name(url):
    '''
    Returns the subfolder name from the insight url.
    Catches the case where an asset sits directly within the root folder.
    '''
    folder = url.rsplit('/')[-2]
    if folder == 's3-ce100-library':
        return ''
    return url.rsplit('/')[-2]

def get_s3_key(url):
    '''
    Returns the s3_key to download a file from the insight url.
    '''
    folder_name = get_folder_name(url)
    if folder_name == '':
        return get_filename(url)
    else:
        return get_folder_name(url) + '/' + get_filename(url)

def get_type_code(type):
    return type_dict[type]

def download_file(s3, bucket_name, url, destination_folder):

    if not destination_folder.endswith('/'):
        destination_folder += '/'

    filename = get_filename(url)
    print('filename ', filename)
    path = destination_folder + filename
    print('bucket: {} s3_key: {} '.format(bucket_name, get_s3_key(url)))

    if not os.path.isfile(path):

        try:
            s3.Bucket(bucket_name).download_file(get_s3_key(url), path)

            print('File {} successfully downloaded \nfrom Bucket: {}, Url: {}'.format(filename, bucket_name, url))

        except botocore.exceptions.ClientError as e:
            if e.response['Error']['Code'] == "404":
                print("The object does not exist. {}".format(get_s3_key(url)))
            else:
                raise

    else:
        print('File {} already exists.'.format(path))

def create_folder(folder_name):

    try:
        folder_parent = Folder.objects.get(name='Cello Library')

        if folder_name == '':
            return folder_parent

        folders = Folder.objects.filter(parent=folder_parent)
        print('current folders: {}'.format(folders))

        folder_exists = False

        for folder in folders:
            if folder.name == folder_name:
                folder_exists = True
                print('Folder {} already exists'.format(folder_name))
                return folder

        if not folder_exists:
            f = Folder(parent=folder_parent, name=folder_name)
            f.save()
            print('created folder {}, parent = {}'.format(f.name, f.parent))
            return f

    except ObjectDoesNotExist as e:
        print('Searching for Cello Library Folder... {}'.format(e))


# def create_asset(name, filename, folder_name, type_field, authors, created_at, active, resource):
def create_asset(insight_details):

    isFile = False
    parent = create_folder('websites')

    if(insight_details['url'].startswith('http://s3')):
        isFile = True
        parent = create_folder(get_folder_name(insight_details['url']))

    # update asset if it exists already
    asset = Asset.objects.filter(parent=parent, name=insight_details['title']).first()
    if asset is None:
        # create
        asset = Asset(
            name=insight_details['title'],
            parent=parent,
        )
        # save must be performed before many-to-many field applied
        asset.save()

    asset.type_field = get_type_code(insight_details['insight_type'])
    asset.contributors = get_or_create_contributors(insight_details['author'])
    asset.uploaded_at = dateutil.parser.parse(insight_details['date'])
    asset.enabled = insight_details['active']
    asset.collections.add(get_collection_from_resource_flag(insight_details['resource']))

    asset.save()

    # pickup if url is a link, not a file!
    if isFile:
        filename = get_filename(insight_details['url'])
        f = open(settings.BASE_DIR + '/file_manager/scripts/ce100_migration/temporary_files/{}'.format(filename), 'rb')

        if len(filename) > 100:
            extension = filename.rsplit('.')[-1]
            print('{} renamed to {}.'.format(filename, filename[:90] + extension))
            filename = filename[:90] + extension

        asset.file.save(filename, File(f))

    else:
        asset.link = insight_details['url']
        asset.save()

    asset_dict[insight_details['insight_id']] = asset

def migrate_asset(insight_details):
    download_file(s3, SOURCE_BUCKET_NAME, insight_details['url'], 'temporary_files')
    create_asset(insight_details)

def read_asset(row):
    '''
    Reads insight details from excel spreadsheet, calling migrate_asset() on each.
    '''
    if insight_sheet.cell(row=row, column=1).value is not None:
        insight_details = {
            'insight_id' : insight_sheet.cell(row=row, column=1).value,
            'title' : insight_sheet.cell(row=row, column=2).value.strip().strip('\n'),
            'url' : insight_sheet.cell(row=row, column=3).value,
            'insight_type' : insight_sheet.cell(row=row, column=4).value,
            'author' : insight_sheet.cell(row=row, column=5).value,
            'date' : insight_sheet.cell(row=row, column=6).value,
        }

        if str(insight_sheet.cell(row=row, column=7).value) == 'True':
            insight_details['active'] = True
        else:
            insight_details['active'] = False

        if str(insight_sheet.cell(row=row, column=8).value) == 'True':
            insight_details['resource'] = True
        else:
            insight_details['resource'] = False

        migrate_asset(insight_details)

        read_asset(row+1)

def read_tag(row):

    if tag_sheet.cell(row=row, column=1).value is not None:

        tag = get_tag(str(tag_sheet.cell(row=row, column=1).value))

        # if tag == -1:
        #     print('ERROR Tag: {} not found.'.format(str(tag_sheet.cell(row=row, column=1).value)))

        if tag == -1:
            print('Attempting to add Tag: "{}" to Insight: {} but tag does not exist... please add manually.'.format(str(tag_sheet.cell(row=row, column=1).value), asset_dict[tag_sheet.cell(row=row, column=2).value]))

        asset = asset_dict[tag_sheet.cell(row=row, column=2).value]

        asset.tags.add(tag)

        print('Added Tag: "{}" to Insight: {}'.format(
            str(tag_sheet.cell(row=row, column=1).value),
            asset_dict[tag_sheet.cell(row=row, column=2).value]
            # str(tag_sheet.cell(row=row, column=2).value),
            )
        )

        read_tag(row+1)

def add_tags_to_assets():

    read_tag(2)

def run():
    print("Well done, ce100 library successfully migrated... well, almost :-)")

    # test read_asset

    # insight_details = {
    #     'insight_id' : str(insight_sheet.cell(row=2, column=1).value),
    #     'title' : insight_sheet.cell(row=2, column=2).value,
    #     'url' : insight_sheet.cell(row=2, column=3).value,
    #     'insight_type' : insight_sheet.cell(row=2, column=4).value,
    #     'author' : insight_sheet.cell(row=2, column=5).value,
    #     'date' : insight_sheet.cell(row=2, column=6).value,
    #     'active' : str(insight_sheet.cell(row=2, column=7).value),
    #     'resource' : str(insight_sheet.cell(row=2, column=8).value),
    # }
    #
    # migrate_asset(insight_details)

    read_asset(2)
    print('Assets added {}'.format(asset_dict))
    add_tags_to_assets()

    # test from niche to norm
    # download_file(s3, SOURCE_BUCKET_NAME, 'http://s3-eu-west-1.amazonaws.com/s3-ce100-library/Circular-furniture-case-studies-1.pdf', 'temporary_files')

    # print('Foo:\n' +
    #     insight_details['insight_id'] + '\n' +
    #     insight_details['title'] + '\n' +
    #     insight_details['url'] + '\n' +
    #     insight_details['insight_type'] + '\n' +
    #     insight_details['author'] + '\n' +
    #     insight_details['date'] + '\n' +
    #     insight_details['active'] + '\n' +
    #     insight_details['resource'] + '\n'
    #     )

    # test get_date
    # print(get_date('2017-07-04T13:44:27.013675+00:00'))

    # test download_file (again!)
    # download_file(s3, SOURCE_BUCKET_NAME, 'http://s3-eu-west-1.amazonaws.com/s3-ce100-library/Insights/A-circular-economy-for-smart-devices.pdf', 'temporary_files')

    # print(get_or_create_contributors('foo, bar2, ray'))
    # print('resource', get_collection_from_resource_flag(True))
    # print('collection', get_collection_from_resource_flag(False))

    # test upload!
    # folder = Folder(name='folder1')
    # folder.save()
    # a = Asset(parent=folder, name='asset1')
    # f = open(settings.BASE_DIR + '/file_manager/scripts/ce100_migration/temporary_files/{}'.format('Agency-of-Design-Case-Study.pdf'), 'rb')
    # myfile = File(f)
    # a.file.save('Agency-of-Design-Case-Study.pdf', myfile)


    # a.save()
    # a.contributors = get_or_create_contributors('George Millard, Alex Wijns')

    # contributors = get_or_create_contributors('George Millard, Alex Wijns')
    # for contributor in contributors:
    #     print(contributor)

    # a.contributors = get_or_create_contributors('George Millard, Alex Wijns')
    # a.save()
