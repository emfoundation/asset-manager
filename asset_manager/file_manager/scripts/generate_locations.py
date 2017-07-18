import openpyxl
import os

from file_manager.models import ContinentTagGroup, CountryTag

os.chdir('file_manager/scripts/location_list/')
wb = openpyxl.load_workbook('country_list.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

CONTINENT_NAMES = {
    'AF':'Africa',
    'AS':'Asia',
    'EU':'Europe',
    'NA':'North America',
    'SA':'South America',
    'OC':'Oceania',
    'AN':'Antarctica',
}

def create_continent_tag(code, name):
    for continent in ContinentTagGroup.objects.all():
        if continent.code == code:
            return 'Continent {} already exists.'.format(name)
    continent = ContinentTagGroup(name=name, code=code)
    continent.save()
    return 'Created continent tag for {}, code: {}'.format(name, code)

def create_country_tag(continent, code, name):
    # when dealing with models check country does not already exist
    for country in CountryTag.objects.all():
        if country.name == name:
            return 'Country {} already exists.'.format(name)
    country = CountryTag(name=name, code=code, continent=ContinentTagGroup.objects.get(name=continent))
    country.save()
    print('Created country tag for {}, code: {}, in continent {}'.format(name, code, continent))

def evaluate_row(row):
    if sheet.cell(row=row, column=1).value is not None:
        continent_code = sheet.cell(row=row, column=1).value
        continent_name = CONTINENT_NAMES[continent_code]
        country_code = sheet.cell(row=row, column=3).value
        country_name = sheet.cell(row=row, column=5).value

        print(create_continent_tag(continent_code, continent_name))
        create_country_tag(continent_name, country_code, country_name)

        evaluate_row(row+1)

def run():
    print(os.getcwd())

    evaluate_row(2)
